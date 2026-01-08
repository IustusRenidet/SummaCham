from __future__ import annotations

import json
from pathlib import Path
import re
import unicodedata

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
EXCEL_PATH = ROOT / 'info IMPORTANTE' / 'CUENTAS RESUMEN.xlsx'
PLANTILLAS_DIR = ROOT / 'PLANTILLAS 2026+'
DESTINO = ROOT / 'vistas' / 'js' / 'resumen-data.js'
AGG_INDICES = [4, 6, 8, 10, 12, 14, 16, 18]


def to_text(value):
    if value is None:
        return ''
    return str(value).strip()


def normalizar(texto):
    valor = to_text(texto)
    if not valor:
        return ''
    nfkd = unicodedata.normalize('NFD', valor)
    limpio = ''.join(ch for ch in nfkd if unicodedata.category(ch) != 'Mn')
    return limpio.upper()


def leer_json(path: Path):
    try:
        return json.loads(path.read_text(encoding='utf-8'))
    except UnicodeDecodeError:
        return json.loads(path.read_text(encoding='latin-1'))


def obtener_valor(item, claves_normalizadas):
    for clave, valor in item.items():
        if normalizar(clave) in claves_normalizadas:
            texto = to_text(valor)
            if texto:
                return texto
    return ''


def extraer_cuentas_desde_plantillas(plantillas_dir: Path):
    if not plantillas_dir.exists():
        return None

    registros = []
    json_paths = sorted(plantillas_dir.rglob('*_layout.json'), key=lambda p: p.name.lower())
    if not json_paths:
        return None

    for json_path in json_paths:
        capitulo_dir = json_path.parent.name.replace(' 2026', '').strip()
        contenido = leer_json(json_path)
        if not isinstance(contenido, dict) or not contenido:
            continue

        for modulo, items in contenido.items():
            if normalizar(modulo) != 'RESUMEN':
                continue
            if not isinstance(items, list):
                items = [items]
            for item in items:
                if not isinstance(item, dict):
                    continue
                cuenta = obtener_valor(item, {'CUENTA'})
                if not cuenta:
                    continue
                capitulo = obtener_valor(item, {'CAPITULO'}) or capitulo_dir
                principal = obtener_valor(item, {'SECCION PRINCIPAL'})
                secundaria = obtener_valor(item, {'SECCION SECUNDARIA'})
                if secundaria:
                    seccion = formatear_seccion(secundaria, principal)
                else:
                    seccion = formatear_seccion(principal, '')
                if not seccion:
                    continue
                registros.append({
                    'capitulo': capitulo,
                    'seccion': seccion,
                    'cuenta': cuenta,
                    'nombre': obtener_valor(item, {'NOMBRE'}) or cuenta
                })

    return registros or None


def leer_resumen_existente(path: Path):
    if not path.exists():
        return {}
    contenido = path.read_text(encoding='utf-8')
    match = re.search(r'const resumenData = (\{.*?\});', contenido, re.S)
    if not match:
        return {}
    try:
        return json.loads(match.group(1))
    except json.JSONDecodeError:
        return {}


def tipo_desde_clase(clase):
    texto = to_text(clase)
    if not texto:
        return ''
    if '-' not in texto:
        return texto.upper()
    return texto.split('-', 1)[0].strip().upper()


def formatear_seccion(base, tipo):
    nombre = to_text(base)
    etiqueta = to_text(tipo).upper()
    if not nombre:
        return etiqueta or ''
    if etiqueta:
        return f'{nombre} ({etiqueta})'
    return nombre


def extraer_cuentas(libro):
    registros = []
    for nombre_hoja in libro.sheetnames:
        if nombre_hoja == 'Filas SUMAS':
            continue
        capitulo = nombre_hoja.replace('Resumen ', '').strip()
        if not capitulo:
            continue
        hoja = libro[nombre_hoja]
        cabecera = True
        for fila in hoja.iter_rows(values_only=True):
            if not fila or not fila[0]:
                cabecera = False
                continue
            cuenta = to_text(fila[0])
            if not cuenta:
                continue
            if cabecera and cuenta.lower().startswith('cuenta'):
                cabecera = False
                continue
            cabecera = False
            descripcion = to_text(fila[1])
            seccion_detallada = to_text(fila[2] if len(fila) > 2 else '')
            seccion_mayor = to_text(fila[3] if len(fila) > 3 else '')
            seccion = formatear_seccion(seccion_detallada or seccion_mayor, seccion_mayor)
            if not seccion:
                continue
            registros.append({
                'capitulo': capitulo,
                'seccion': seccion,
                'cuenta': cuenta,
                'nombre': descripcion
            })
    return registros


def extraer_sumas(libro):
    if 'Filas SUMAS' not in libro.sheetnames:
        return {}
    hoja = libro['Filas SUMAS']
    sumas = {}
    primera = True
    for fila in hoja.iter_rows(values_only=True):
        if primera:
            primera = False
            continue
        if not fila:
            continue
        capitulo = to_text(fila[0])
        seccion = to_text(fila[2])
        if not capitulo or not seccion:
            continue
        etiquetas = []
        for idx in AGG_INDICES:
            if idx >= len(fila):
                continue
            valor = fila[idx]
            if valor is None:
                continue
            texto = to_text(valor)
            if texto:
                etiquetas.append(texto)
        sumRowSumavarios = etiquetas[0] if etiquetas else ''
        sumRowSumavarios2 = etiquetas[1] if len(etiquetas) > 1 else ''
        resultRows = etiquetas[2:]
        clase = to_text(fila[1])
        tipo = tipo_desde_clase(clase)
        seccion_label = formatear_seccion(seccion, tipo)
        if not seccion_label:
            continue
        sumas.setdefault(capitulo, {})[seccion_label] = {
            'sumRow': '',
            'sumRowSumavarios': sumRowSumavarios,
            'sumRowSumavarios2': sumRowSumavarios2,
            'resultRows': [texto for texto in resultRows if texto]
        }
    return sumas


def main():
    libro = None
    if EXCEL_PATH.exists():
        libro = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    cuentas = extraer_cuentas_desde_plantillas(PLANTILLAS_DIR)
    if cuentas is None and libro is not None:
        cuentas = extraer_cuentas(libro)
    if cuentas is None:
        cuentas = leer_resumen_existente(DESTINO).get('cuentas', [])

    if libro is not None:
        sumas = extraer_sumas(libro)
    else:
        sumas = leer_resumen_existente(DESTINO).get('sumas', {})
    resumen_data = {'cuentas': cuentas, 'sumas': sumas}
    contenido = (
        '// Generado automaticamente desde PLANTILLAS 2026+ (cuentas) y info IMPORTANTE/CUENTAS RESUMEN.xlsx (sumas)\n'
        '(function () {\n'
        f'  const resumenData = {json.dumps(resumen_data, ensure_ascii=False, indent=2)};\n'
        '  window.CUENTAS_POR_MODULO = window.CUENTAS_POR_MODULO || {};\n'
        "  window.CUENTAS_POR_MODULO['resumen'] = resumenData.cuentas;\n"
        '  window.CUENTAS_SUMAS = window.CUENTAS_SUMAS || {};\n'
        "  window.CUENTAS_SUMAS['RESUMEN'] = resumenData.sumas;\n"
        '})();\n'
    )
    DESTINO.write_text(contenido, encoding='utf-8')
    print(f'Archivo actualizado: {DESTINO}')


if __name__ == '__main__':
    main()
