#!/usr/bin/env python3
import argparse
import json
import os
import re
import sys
import unicodedata

try:
    from openpyxl import load_workbook
    from openpyxl.chart import BarChart, DoughnutChart, LineChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.utils import get_column_letter
except Exception as exc:  # pragma: no cover - runtime dependency guard
    print(str(exc), file=sys.stderr)
    raise

CHART_GAP_ROWS = 1


def text(v):
    if v is None:
        return ""
    raw = str(v).strip()
    if not raw:
        return ""
    cleaned = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F\x7F-\x9F]", " ", raw)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned


def normalized_text(v):
    raw = text(v)
    if not raw:
        return ""
    decomposed = unicodedata.normalize("NFD", raw)
    stripped = "".join(ch for ch in decomposed if unicodedata.category(ch) != "Mn")
    cleaned = re.sub(r"[^A-Za-z0-9]+", " ", stripped).strip()
    return cleaned.upper()


def _clamp_byte(value):
    try:
        return max(0, min(255, int(round(float(value)))))
    except Exception:
        return 0


def _parse_percent_or_number(raw, max_value=1.0):
    if raw is None:
        return None
    txt = str(raw).strip()
    if not txt:
        return None
    try:
        if txt.endswith("%"):
            pct = float(txt[:-1])
            return (pct / 100.0) * max_value
        return float(txt)
    except Exception:
        return None


def _blend_rgba_to_hex(r, g, b, a=1.0):
    alpha_raw = _parse_percent_or_number(a, 1.0)
    alpha = alpha_raw if alpha_raw is not None else 1.0
    alpha = max(0.0, min(1.0, float(alpha)))

    def blend(channel):
        base = _clamp_byte(channel)
        return int(round(alpha * base + (1.0 - alpha) * 255.0))

    rb = blend(r)
    gb = blend(g)
    bb = blend(b)
    return f"{rb:02X}{gb:02X}{bb:02X}"


def sanitize_hex(color, fallback="4472C4"):
    source_raw = text(color).strip()
    source = source_raw.lstrip("#")

    if re.fullmatch(r"[0-9A-Fa-f]{3}", source):
        return "".join(ch * 2 for ch in source).upper()
    if re.fullmatch(r"[0-9A-Fa-f]{6}", source):
        return source.upper()
    if re.fullmatch(r"[0-9A-Fa-f]{4}", source):
        expanded = "".join(ch * 2 for ch in source)
        r = int(expanded[0:2], 16)
        g = int(expanded[2:4], 16)
        b = int(expanded[4:6], 16)
        a = int(expanded[6:8], 16) / 255.0
        return _blend_rgba_to_hex(r, g, b, a)
    if re.fullmatch(r"[0-9A-Fa-f]{8}", source):
        r = int(source[0:2], 16)
        g = int(source[2:4], 16)
        b = int(source[4:6], 16)
        a = int(source[6:8], 16) / 255.0
        return _blend_rgba_to_hex(r, g, b, a)

    rgb_match = re.fullmatch(
        r"rgba?\(\s*([0-9]{1,3}%?)\s*[,\s]\s*([0-9]{1,3}%?)\s*[,\s]\s*([0-9]{1,3}%?)(?:\s*[,/]\s*([0-9.]+%?))?\s*\)",
        source_raw,
        re.IGNORECASE,
    )
    if rgb_match:
        def parse_channel(comp):
            parsed = _parse_percent_or_number(comp, 255.0)
            return _clamp_byte(parsed if parsed is not None else 0)

        r = parse_channel(rgb_match.group(1))
        g = parse_channel(rgb_match.group(2))
        b = parse_channel(rgb_match.group(3))
        a = rgb_match.group(4) if rgb_match.group(4) is not None else 1.0
        return _blend_rgba_to_hex(r, g, b, a)

    return fallback


def clear_sheet(ws):
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    if ws.max_column > 0:
        ws.delete_cols(1, ws.max_column)
    ws._charts = []


def clear_charts_only(ws):
    ws._charts = []


def best_table_sheet(wb, preferred_name):
    if preferred_name and preferred_name in wb.sheetnames:
        return wb[preferred_name]
    if wb.sheetnames:
        return wb[wb.sheetnames[0]]
    return None


def autofit_columns(ws, min_width=10, max_width=60):
    if ws is None:
        return
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            candidate = len(str(v))
            if candidate > max_len:
                max_len = candidate
        width = max(min_width, min(max_width, max_len + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def parse_series_meta(series_meta):
    if not series_meta:
        return []
    try:
        parsed = json.loads(series_meta)
        if isinstance(parsed, list):
            return parsed
        if isinstance(parsed, dict):
            return [parsed]
    except Exception:
        return []
    return []


def _entry_series_key(entry):
    if not isinstance(entry, dict):
        return ""
    return normalized_text(entry.get("key") or entry.get("label") or entry.get("name") or "")


def _entry_chart_key(entry):
    if not isinstance(entry, dict):
        return ""
    return normalized_text(entry.get("chartKey") or entry.get("chartTitle") or entry.get("title") or "")


def _entry_order(entry):
    if not isinstance(entry, dict):
        return None
    try:
        parsed = int(entry.get("order"))
        return parsed
    except Exception:
        return None


def find_meta_entry(meta_list, name, idx, chart_title=""):
    target = normalized_text(name)
    chart_key = normalized_text(chart_title)
    entries = [entry for entry in meta_list if isinstance(entry, dict)]

    # 1) Match exact by chart + series key.
    if chart_key:
        for entry in entries:
            if _entry_chart_key(entry) == chart_key and _entry_series_key(entry) == target and target:
                return entry

    # 2) Match global by series key.
    for entry in entries:
        if _entry_series_key(entry) == target and target:
            return entry

    # 3) Match by chart + order as fallback.
    if chart_key:
        for entry in entries:
            if _entry_chart_key(entry) == chart_key and _entry_order(entry) == idx:
                return entry

    # 4) Match global by order.
    for entry in entries:
        if _entry_order(entry) == idx:
            return entry

    return None


def resolve_series_type(meta_list, name, idx, chart_title=""):
    entry = find_meta_entry(meta_list, name, idx, chart_title)
    if entry:
        raw = text(entry.get("type") or entry.get("chartType")).lower()
        if raw in {"line", "bar", "pie", "doughnut"}:
            return raw
    return "bar"


def _entry_orientation(entry):
    if not isinstance(entry, dict):
        return ""
    orientation = text(entry.get("chartOrientation") or entry.get("orientation")).lower()
    if orientation in {"horizontal", "vertical"}:
        return orientation
    index_axis = text(entry.get("indexAxis") or entry.get("chartIndexAxis")).lower()
    if index_axis == "y":
        return "horizontal"
    if index_axis == "x":
        return "vertical"
    return ""


def _resolve_chart_orientation(meta_list, iterable, chart_title=""):
    entries = _collect_chart_meta_entries(meta_list, iterable, chart_title)
    for entry in entries:
        orientation = _entry_orientation(entry)
        if orientation:
            return orientation
    return "vertical"


def _collect_chart_meta_entries(meta_list, iterable, chart_title=""):
    entries = []
    for idx, (_, series_name) in enumerate(iterable):
        entry = find_meta_entry(meta_list, series_name, idx, chart_title)
        if entry:
            entries.append(entry)
    chart_key = normalized_text(chart_title)
    if chart_key:
        for entry in meta_list:
            if isinstance(entry, dict) and _entry_chart_key(entry) == chart_key:
                entries.append(entry)
    return entries


def _resolve_chart_visual_type(meta_list, iterable, chart_title=""):
    series_types = []
    for idx, (_, series_name) in enumerate(iterable):
        series_types.append(resolve_series_type(meta_list, series_name, idx, chart_title))
    effective = [t for t in series_types if t]
    if not effective:
        return "cartesian"
    if all(t in {"pie", "doughnut"} for t in effective):
        return "doughnut" if "doughnut" in effective else "pie"
    return "cartesian"


def resolve_series_colors(meta_list, name, idx, chart_title="", series_type="bar"):
    entry = find_meta_entry(meta_list, name, idx, chart_title)
    if entry:
        base_raw = text(entry.get("color"))
        fill_raw = text(entry.get("fillColor") or entry.get("fill") or base_raw)
        line_raw = text(entry.get("lineColor") or entry.get("strokeColor") or base_raw or fill_raw)
        base = sanitize_hex(base_raw, "4472C4") if base_raw else "4472C4"
        fill = sanitize_hex(fill_raw, base)
        line = sanitize_hex(line_raw, fill)
        primary = line if series_type == "line" else fill
        return fill, line, primary

    norm = normalized_text(name)
    if "NET" in norm or "NETO" in norm:
        default = "94A3B8"
        return default, default, default
    if "ANTERIOR" in norm or "AA" in norm or "PREV" in norm:
        default = "94A3B8"
        return default, default, default
    if "PPTO" in norm or "PRESUPUESTO" in norm or "BUDGET" in norm:
        default = "60A5FA"
        return default, default, default
    if "REAL" in norm:
        default = "0D47A1"
        return default, default, default
    palette = ["0D47A1", "60A5FA", "94A3B8", "F59E0B", "10B981", "4472C4"]
    default = palette[idx % len(palette)]
    return default, default, default


def style_series(series, fill_hex, line_hex, series_type):
    fill_color = sanitize_hex(fill_hex, "4472C4")
    line_color = sanitize_hex(line_hex, fill_color)
    if series_type == "bar":
        try:
            # Mantener el mismo color para valores negativos (solo barras).
            series.invertIfNegative = False
        except Exception:
            pass
    try:
        series.graphicalProperties.solidFill = fill_color
    except Exception:
        pass
    if series_type in {"bar", "line"}:
        try:
            series.graphicalProperties.line.solidFill = line_color
        except Exception:
            pass
    if series_type == "line":
        # Configuración conservadora para evitar XML inválido en ciertas versiones de Excel.
        # Evitamos tocar marker/marker properties y solo aplicamos color de línea.
        pass


def enable_value_labels(chart, show_percent=False, show_values=True, dense=False):
    try:
        labels = DataLabelList()
        labels.showVal = bool(show_values)
        labels.showPercent = bool(show_percent)
        labels.showSerName = False
        labels.showCatName = False
        labels.showLegendKey = False
        labels.showBubbleSize = False
        labels.showLeaderLines = bool(show_percent and not dense)
        chart.dataLabels = labels
    except Exception:
        pass


def _chart_rows_from_height(height_inches):
    try:
        height = float(height_inches)
    except Exception:
        height = 8.0
    # Compactar separación vertical entre gráficas sin encimarlas.
    return max(15, int(round(height * 4.4)) + CHART_GAP_ROWS)


def _max_category_label_len(ws_data, block):
    if ws_data is None or not isinstance(block, dict):
        return 0
    start = int(block.get("data_start") or 0)
    end = int(block.get("data_end") or 0)
    if start <= 0 or end <= 0 or end < start:
        return 0
    max_len = 0
    for row in range(start, end + 1):
        label = text(ws_data.cell(row=row, column=1).value)
        if not label:
            continue
        max_len = max(max_len, len(label))
    return max_len


def _compute_chart_size(
    labels_count,
    series_count,
    orientation="vertical",
    is_combined=False,
    chart_visual_type="cartesian",
    max_label_len=0,
):
    labels_count = max(1, int(labels_count or 1))
    series_count = max(1, int(series_count or 1))
    orientation = "horizontal" if orientation == "horizontal" else "vertical"
    max_label_len = max(0, int(max_label_len or 0))

    if is_combined:
        width = (
            31.0
            if orientation == "horizontal"
            else max(25.0, min(36.0, 23.0 + max_label_len * 0.18))
        )
        if orientation == "horizontal":
            height = max(
                11.0,
                min(32.0, 8.8 + labels_count * 0.54 + (1.6 if max_label_len > 24 else 0.0)),
            )
        else:
            height = max(
                8.2,
                min(19.0, 6.2 + labels_count * 0.19 + min(2.8, max_label_len * 0.018)),
            )
        return width, height

    if chart_visual_type in {"pie", "doughnut"}:
        width = 20.0
        height = max(7.4, min(12.8, 6.2 + labels_count * 0.14))
        return width, height

    width = (
        max(30.0, min(44.0, 30.0 + max_label_len * 0.18))
        if orientation == "horizontal"
        else max(24.0, min(38.0, 24.0 + max_label_len * 0.12))
    )
    if orientation == "horizontal":
        height = max(
            9.6,
            min(
                25.0,
                6.8
                + labels_count * 0.50
                + (series_count - 1) * 0.16
                + (1.2 if max_label_len > 24 else 0.0),
            ),
        )
    else:
        height = max(
            7.6,
            min(
                16.8,
                5.8
                + labels_count * 0.13
                + (series_count - 1) * 0.08
                + (0.8 if max_label_len > 22 else 0.0),
            ),
        )
    return width, height


def parse_resumen_blocks(ws_data):
    blocks = []
    max_row = ws_data.max_row
    max_col = ws_data.max_column
    row = 1
    while row <= max_row:
        marker = normalized_text(ws_data.cell(row=row, column=1).value)
        if marker != "CHART":
            row += 1
            continue
        title = text(ws_data.cell(row=row, column=2).value) or f"Grafica {len(blocks) + 1}"
        header_row = row + 1
        if header_row > max_row:
            break

        col = 2
        while col <= max_col and text(ws_data.cell(row=header_row, column=col).value):
            col += 1
        last_col = col - 1

        data_start = header_row + 1
        data_end = data_start - 1
        while data_start <= max_row and text(ws_data.cell(row=data_start, column=1).value):
            data_end = data_start
            data_start += 1

        if last_col >= 2 and data_end >= header_row + 1:
            blocks.append(
                {
                    "title": title,
                    "header_row": header_row,
                    "data_start": header_row + 1,
                    "data_end": data_end,
                    "series_cols": list(range(2, last_col + 1)),
                }
            )
            row = data_end + 2
        else:
            row += 1
    return blocks


def get_series_columns_for_row(ws_data, row_number):
    max_col = ws_data.max_column
    last_col = 1
    for col in range(max_col, 1, -1):
        if text(ws_data.cell(row=row_number, column=col).value):
            last_col = col
            break
    if last_col < 2:
        return []

    columns = []
    for col in range(2, last_col + 1):
        series_name = text(ws_data.cell(row=row_number, column=col).value)
        if not series_name:
            continue
        columns.append({"col": col, "name": series_name})
    return columns


def detect_operativo_header(ws_data):
    max_row = ws_data.max_row
    max_scan = min(120, max_row)
    max_col = ws_data.max_column

    for row in range(1, max_scan + 1):
        parts = []
        for col in range(1, min(12, max_col) + 1):
            candidate = text(ws_data.cell(row=row, column=col).value).lower()
            if candidate:
                parts.append(candidate)
        row_text = " ".join(parts)
        if (
            ("ppto" in row_text or "presupuesto" in row_text or "real" in row_text)
            and "acum" in row_text
        ):
            return row

    metadata_rows = {
        "RESULTADOS OPERATIVOS",
        "CATEGORIA",
        "EMPRESA",
        "PERIODO",
        "FECHA EXPORTACION",
    }

    for row in range(1, max_scan + 1):
        series = get_series_columns_for_row(ws_data, row)
        if not series:
            continue
        first_norm = normalized_text(ws_data.cell(row=row, column=1).value)
        if first_norm in metadata_rows:
            continue
        row_join = []
        for col in range(1, min(12, max_col) + 1):
            v = text(ws_data.cell(row=row, column=col).value)
            if v:
                row_join.append(v)
        row_normalized = normalized_text(" ".join(row_join))
        looks_like_header = bool(
            re.search(r"PPTO|PRESUPUESTO|REAL|ACUM|BUDGET|ACTUAL|YTD|20[0-9]{2}", row_normalized)
        )
        next_label = text(ws_data.cell(row=row + 1, column=1).value) if row + 1 <= max_row else ""
        if looks_like_header or next_label:
            return row
    return 1


def parse_operativo_chart_blocks(ws_data):
    blocks = []
    max_row = ws_data.max_row
    row = 1
    while row <= max_row:
        marker = normalized_text(ws_data.cell(row=row, column=1).value)
        if marker != "CHART":
            row += 1
            continue

        title = text(ws_data.cell(row=row, column=2).value) or f"Grafica {len(blocks) + 1}"
        header_row = row + 1
        if header_row > max_row:
            break

        series = get_series_columns_for_row(ws_data, header_row)
        if not series:
            row = header_row + 1
            continue

        data_start = header_row + 1
        while data_start <= max_row:
            lbl = text(ws_data.cell(row=data_start, column=1).value)
            if not lbl:
                data_start += 1
                continue
            if normalized_text(lbl) == "CHART":
                break
            break
        if data_start > max_row:
            break

        data_end = data_start
        while data_end <= max_row:
            current = text(ws_data.cell(row=data_end, column=1).value)
            if not current:
                break
            if normalized_text(current) == "CHART":
                break
            data_end += 1
        data_end -= 1

        if data_end >= data_start:
            blocks.append(
                {
                    "title": title,
                    "header_row": header_row,
                    "data_start": data_start,
                    "data_end": data_end,
                    "series": series,
                }
            )
            row = data_end + 1
            continue
        row = header_row + 1
    return blocks


def anchor_start_row(ws_data, ws_charts):
    if ws_data is ws_charts:
        return ws_data.max_row + 2
    return 2


def _configure_cartesian_axes(chart, orientation, labels_count=1, max_label_len=0):
    orientation = "horizontal" if orientation == "horizontal" else "vertical"
    labels_count = max(1, int(labels_count or 1))
    max_label_len = max(0, int(max_label_len or 0))
    label_skip = 1
    if labels_count > 36:
        label_skip = 4
    elif labels_count > 26:
        label_skip = 3
    elif labels_count > 22:
        label_skip = 2
    elif labels_count > 18 and max_label_len > 12:
        label_skip = 2
    try:
        if chart.x_axis is not None:
            chart.x_axis.tickLblPos = "low"
    except Exception:
        pass
    try:
        if chart.x_axis is not None:
            chart.x_axis.tickLblSkip = label_skip
    except Exception:
        pass
    try:
        if chart.y_axis is not None:
            chart.y_axis.tickLblSkip = label_skip
    except Exception:
        pass
    try:
        if chart.x_axis is not None:
            chart.x_axis.delete = False
    except Exception:
        pass
    try:
        if chart.y_axis is not None:
            chart.y_axis.delete = False
    except Exception:
        pass
    if orientation == "vertical":
        try:
            # Mantener etiquetas de categorías abajo (evita que "desaparezcan" al cruzar en cero).
            chart.y_axis.crosses = "min"
        except Exception:
            pass
        try:
            if chart.x_axis is not None:
                chart.x_axis.tickLblPos = "low"
        except Exception:
            pass
    else:
        try:
            chart.x_axis.crosses = "autoZero"
        except Exception:
            pass
        try:
            if chart.y_axis is not None:
                chart.y_axis.tickLblPos = "nextTo"
        except Exception:
            pass


def add_chart_for_block(ws_data, ws_charts, block, meta_list, top_row, is_combined=False):
    categories = Reference(
        ws_data, min_col=1, min_row=block["data_start"], max_row=block["data_end"]
    )

    if block.get("series"):
        iterable = [(s["col"], s["name"]) for s in block["series"]]
    else:
        iterable = []
        for col in block.get("series_cols", []):
            iterable.append((col, text(ws_data.cell(row=block["header_row"], column=col).value)))

    chart_title = block.get("title") or "Grafica"
    labels_count = max(1, block["data_end"] - block["data_start"] + 1)
    max_label_len = _max_category_label_len(ws_data, block)
    orientation = _resolve_chart_orientation(meta_list, iterable, chart_title)
    chart_visual_type = _resolve_chart_visual_type(meta_list, iterable, chart_title)

    if chart_visual_type in {"pie", "doughnut"}:
        pie_series = None
        pie_series_name = ""
        for idx, (col_idx, series_name) in enumerate(iterable):
            series_name = series_name or f"Serie {idx + 1}"
            ref = Reference(
                ws_data,
                min_col=col_idx,
                max_col=col_idx,
                min_row=block["header_row"],
                max_row=block["data_end"],
            )
            pie_series = ref
            pie_series_name = series_name
            break

        if pie_series is None:
            return top_row

        base_chart = DoughnutChart() if chart_visual_type == "doughnut" else PieChart()
        base_chart.title = chart_title
        try:
            if chart_visual_type == "doughnut":
                base_chart.holeSize = 55
        except Exception:
            pass
        base_chart.add_data(pie_series, titles_from_data=True)
        base_chart.set_categories(categories)
        show_labels = labels_count <= 12
        enable_value_labels(base_chart, show_percent=show_labels, show_values=False, dense=not show_labels)
        try:
            base_chart.legend.position = "r" if labels_count > 9 else "b"
        except Exception:
            pass

        fill_color, line_color, _ = resolve_series_colors(
            meta_list, pie_series_name, 0, chart_title, chart_visual_type
        )
        if base_chart.series:
            style_series(base_chart.series[0], fill_color, line_color, "bar")

        width, height = _compute_chart_size(
            labels_count,
            1,
            orientation=orientation,
            is_combined=is_combined,
            chart_visual_type=chart_visual_type,
            max_label_len=max_label_len,
        )
        base_chart.width = width
        base_chart.height = height
        row_step = _chart_rows_from_height(height)

        ws_charts.add_chart(base_chart, f"A{top_row}")
        return top_row + row_step

    series_specs = []
    has_line = False
    has_bar = False
    for idx, (col_idx, series_name) in enumerate(iterable):
        series_name = series_name or f"Serie {idx + 1}"
        ref = Reference(
            ws_data,
            min_col=col_idx,
            max_col=col_idx,
            min_row=block["header_row"],
            max_row=block["data_end"],
        )
        requested_type = resolve_series_type(meta_list, series_name, idx, chart_title)
        if requested_type not in {"line", "bar"}:
            requested_type = "bar"
        fill_color, line_color, _ = resolve_series_colors(
            meta_list, series_name, idx, chart_title, requested_type
        )
        has_line = has_line or requested_type == "line"
        has_bar = has_bar or requested_type == "bar"
        series_specs.append(
            {
                "ref": ref,
                "fill": fill_color,
                "line": line_color,
                "requested_type": requested_type,
            }
        )

    if not series_specs:
        return top_row

    # Modo seguro: evitar combinaciones bar+line (axis secundarios),
    # porque en algunos entornos Excel termina reparando el archivo.
    mixed_types = has_line and has_bar
    chart_kind = "line" if has_line and not has_bar else "bar"
    if mixed_types:
        chart_kind = "bar"

    total_series = len(series_specs)
    legend_position = (
        "r"
        if (
            total_series >= 7
            or (total_series >= 5 and labels_count >= 20)
            or max_label_len >= 26
        )
        else "b"
    )

    if chart_kind == "line":
        base_chart = LineChart()
        base_chart.title = chart_title
        base_chart.y_axis.number_format = "#,##0.00"
        for spec in series_specs:
            base_chart.add_data(spec["ref"], titles_from_data=True)
            style_series(base_chart.series[-1], spec["fill"], spec["line"], "line")
        base_chart.set_categories(categories)
    else:
        base_chart = BarChart()
        base_chart.type = "bar" if orientation == "horizontal" else "col"
        base_chart.grouping = "clustered"
        base_chart.overlap = 0
        base_chart.title = chart_title
        try:
            base_chart.varyColors = False
        except Exception:
            pass
        # En openpyxl el eje de valores del BarChart es y_axis en ambos barDir (col/bar).
        base_chart.y_axis.number_format = "#,##0.00"
        for spec in series_specs:
            base_chart.add_data(spec["ref"], titles_from_data=True)
            style_series(base_chart.series[-1], spec["fill"], spec["line"], "bar")
        base_chart.set_categories(categories)

    _configure_cartesian_axes(
        base_chart,
        orientation,
        labels_count=labels_count,
        max_label_len=max_label_len,
    )

    try:
        base_chart.legend.position = legend_position
    except Exception:
        pass

    width, height = _compute_chart_size(
        labels_count,
        total_series,
        orientation=orientation,
        is_combined=is_combined,
        chart_visual_type="cartesian",
        max_label_len=max_label_len,
    )
    base_chart.width = width
    base_chart.height = height
    row_step = _chart_rows_from_height(height)

    ws_charts.add_chart(base_chart, f"A{top_row}")
    return top_row + row_step


def process_resumen(wb, ws_data, ws_charts, table_sheet_name, series_meta):
    blocks = parse_resumen_blocks(ws_data)
    if not blocks:
        raise RuntimeError(f"No chart blocks found in {ws_data.title}.")
    top_row = anchor_start_row(ws_data, ws_charts)
    meta_list = parse_series_meta(series_meta)
    for block in blocks:
        top_row = add_chart_for_block(ws_data, ws_charts, block, meta_list, top_row)

    ws_table = best_table_sheet(wb, table_sheet_name)
    autofit_columns(ws_table)
    if ws_table is not None:
        wb.active = wb.sheetnames.index(ws_table.title)
    elif ws_charts is not None:
        wb.active = wb.sheetnames.index(ws_charts.title)


def process_operativo(wb, ws_data, ws_charts, table_sheet_name, chart_mode, series_meta):
    meta_list = parse_series_meta(series_meta)
    mode = text(chart_mode).lower() or "split"
    top_row = anchor_start_row(ws_data, ws_charts)

    if mode == "combined":
        header_row = detect_operativo_header(ws_data)
        series = get_series_columns_for_row(ws_data, header_row)
        if not series:
            raise RuntimeError(f"No series columns found in {ws_data.title}.")
        data_start = header_row + 1
        while data_start <= ws_data.max_row and not text(ws_data.cell(row=data_start, column=1).value):
            data_start += 1
        if data_start > ws_data.max_row:
            raise RuntimeError(f"No data rows found in {ws_data.title}.")
        block = {
            "title": "Resultados operativos",
            "header_row": header_row,
            "data_start": data_start,
            "data_end": ws_data.max_row,
            "series": series,
        }
        top_row = add_chart_for_block(
            ws_data, ws_charts, block, meta_list, top_row, is_combined=True
        )
    else:
        blocks = parse_operativo_chart_blocks(ws_data)
        if blocks:
            for block in blocks:
                top_row = add_chart_for_block(ws_data, ws_charts, block, meta_list, top_row)
        else:
            header_row = detect_operativo_header(ws_data)
            series = get_series_columns_for_row(ws_data, header_row)
            if not series:
                raise RuntimeError(f"No series columns found in {ws_data.title}.")
            data_start = header_row + 1
            while data_start <= ws_data.max_row and not text(ws_data.cell(row=data_start, column=1).value):
                data_start += 1
            if data_start > ws_data.max_row:
                raise RuntimeError(f"No data rows found in {ws_data.title}.")
            for idx, serie in enumerate(series):
                block = {
                    "title": serie["name"] or f"Serie {idx + 1}",
                    "header_row": header_row,
                    "data_start": data_start,
                    "data_end": ws_data.max_row,
                    "series": [serie],
                }
                top_row = add_chart_for_block(ws_data, ws_charts, block, meta_list, top_row)

    ws_table = best_table_sheet(wb, table_sheet_name)
    autofit_columns(ws_table)
    if ws_table is not None:
        wb.active = wb.sheetnames.index(ws_table.title)
    elif ws_charts is not None:
        wb.active = wb.sheetnames.index(ws_charts.title)


def run():
    parser = argparse.ArgumentParser()
    parser.add_argument("--kind", default="operativo")
    parser.add_argument("--input-path", required=True)
    parser.add_argument("--output-path", required=True)
    parser.add_argument("--data-sheet-name", default="GraficasData")
    parser.add_argument("--charts-sheet-name", default="Graficas")
    parser.add_argument("--table-sheet-name", default="")
    parser.add_argument("--chart-mode", default="split")
    parser.add_argument("--series-meta", default="")
    args = parser.parse_args()

    if not os.path.exists(args.input_path):
        raise RuntimeError(f"Input file not found: {args.input_path}")

    wb = load_workbook(args.input_path)

    if args.data_sheet_name not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise RuntimeError(
            f"Data sheet not found: {args.data_sheet_name}. Available: {available}"
        )
    ws_data = wb[args.data_sheet_name]

    if args.charts_sheet_name:
        if args.charts_sheet_name in wb.sheetnames:
            ws_charts = wb[args.charts_sheet_name]
            if ws_charts is ws_data:
                clear_charts_only(ws_charts)
            else:
                clear_sheet(ws_charts)
        else:
            ws_charts = wb.create_sheet(args.charts_sheet_name)
    else:
        ws_charts = ws_data
        clear_charts_only(ws_charts)

    kind = text(args.kind).lower()
    if kind == "resumen":
        process_resumen(wb, ws_data, ws_charts, args.table_sheet_name, args.series_meta)
    else:
        process_operativo(
            wb, ws_data, ws_charts, args.table_sheet_name, args.chart_mode, args.series_meta
        )

    os.makedirs(os.path.dirname(os.path.abspath(args.output_path)), exist_ok=True)
    wb.save(args.output_path)
    print(f"Charts created: {args.output_path}")


if __name__ == "__main__":
    try:
        run()
    except Exception as exc:
        print(str(exc), file=sys.stderr)
        sys.exit(1)
