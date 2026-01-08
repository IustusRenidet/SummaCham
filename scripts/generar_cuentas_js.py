from __future__ import annotations

import json
from pathlib import Path
import unicodedata

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
EXCEL_PATH = ROOT / 'info IMPORTANTE' / 'CUENTAS.xlsx'
PLANTILLAS_DIR = ROOT / 'PLANTILLAS 2026+'
DESTINO = ROOT / 'vistas' / 'js' / 'cuentas-data.js'
HOJA_SUMAS = 'SUMA DE VARIAS SECCIONES'


def normalizar(texto):
  valor = (texto or '').strip()
  if not valor:
    return ''
  nfkd = unicodedata.normalize('NFD', valor)
  limpio = ''.join(ch for ch in nfkd if unicodedata.category(ch) != 'Mn')
  return limpio.upper()


def normalizar_hoja(texto):
  limpio = normalizar(texto)
  return ''.join(ch for ch in limpio if ch not in {'.', ' ', '_'})


def leer_json(path: Path):
  try:
    return json.loads(path.read_text(encoding='utf-8'))
  except UnicodeDecodeError:
    return json.loads(path.read_text(encoding='latin-1'))


def limpiar_modulo(texto):
  base = (texto or '').strip()
  if not base:
    return ''
  return ' '.join(base.replace('.', ' ').replace('_', ' ').split())


def obtener_valor(item, claves_normalizadas):
  for clave, valor in item.items():
    if normalizar(clave) in claves_normalizadas:
      texto = (valor or '').strip() if isinstance(valor, str) else str(valor or '').strip()
      if texto:
        return texto
  return ''


def es_modulo_resumen(nombre):
  return normalizar(nombre) in {'RESUMEN', 'SUMMARY'}


def extraer_registros_desde_plantillas(plantillas_dir: Path):
  if not plantillas_dir.exists():
    return None

  datos = {}
  json_paths = sorted(plantillas_dir.rglob('*_layout.json'), key=lambda p: p.name.lower())
  if not json_paths:
    return None

  for json_path in json_paths:
    capitulo_dir = json_path.parent.name.replace(' 2026', '').strip()
    contenido = leer_json(json_path)
    if not isinstance(contenido, dict) or not contenido:
      continue

    for modulo, items in contenido.items():
      if es_modulo_resumen(modulo):
        continue
      modulo_limpio = limpiar_modulo(modulo)
      if not modulo_limpio:
        continue
      if not isinstance(items, list):
        items = [items]
      for item in items:
        if not isinstance(item, dict):
          continue
        cuenta = obtener_valor(item, {'CUENTA'})
        if not cuenta:
          continue
        capitulo = obtener_valor(item, {'CAPITULO'}) or capitulo_dir
        seccion = obtener_valor(item, {'SECCION', 'SECCION PRINCIPAL'})
        nombre = obtener_valor(item, {'NOMBRE'}) or cuenta
        datos.setdefault(modulo_limpio, []).append({
          'capitulo': capitulo,
          'seccion': seccion,
          'cuenta': cuenta,
          'nombre': nombre
        })

  return datos or None


def extraer_registros_excel(libro):
  datos = {}
  for hoja in libro.worksheets:
    if hoja.title == HOJA_SUMAS:
      continue
    registros = []
    for fila in hoja.iter_rows(values_only=True):
      if not fila:
        continue
      capitulo, seccion, cuenta, nombre = (fila + (None, None, None, None))[:4]
      if capitulo == 'CAPITULO' or not (capitulo and cuenta):
        continue
      registros.append({
        'capitulo': str(capitulo).strip(),
        'seccion': str(seccion or '').strip(),
        'cuenta': str(cuenta or '').strip(),
        'nombre': str(nombre or '').strip()
      })
    datos[hoja.title] = registros
  return datos


def extraer_sumas(libro):
  hoja = libro[HOJA_SUMAS]
  sumas = {}
  encabezado = True
  for fila in hoja.iter_rows(values_only=True):
    if encabezado:
      encabezado = False
      continue
    if not fila:
      continue
    hoja_id, capitulo, seccion, sum_row, sum_row_suma, sum_row_suma2, result_row = (fila + (None,) * 7)[:7]
    if not hoja_id or not capitulo or not seccion:
      continue
    hoja_key = normalizar_hoja(str(hoja_id))
    cap_key = normalizar(capitulo)
    sec_key = normalizar(seccion)
    sumas.setdefault(hoja_key, {}).setdefault(cap_key, {})[sec_key] = {
      'sumRow': (sum_row or '').strip(),
      'sumRowSumavarios': (sum_row_suma or '').strip(),
      'sumRowSumavarios2': (sum_row_suma2 or '').strip(),
      'resultRow': (result_row or '').strip()
    }
  return sumas


def main():
  libro = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
  registros = extraer_registros_desde_plantillas(PLANTILLAS_DIR)
  if registros is None:
    registros = extraer_registros_excel(libro)
  sumas = extraer_sumas(libro)
  contenido = (
    '// Generado automaticamente desde PLANTILLAS 2026+ (cuentas) y info IMPORTANTE/CUENTAS.xlsx (sumas)\n'
    f'window.CUENTAS_POR_MODULO = {json.dumps(registros, ensure_ascii=True, separators=(",", ":"))};\n'
    f'window.CUENTAS_SUMAS = {json.dumps(sumas, ensure_ascii=True, separators=(",", ":"))};\n'
  )
  DESTINO.write_text(contenido, encoding='utf-8')
  print(f'Archivo actualizado: {DESTINO}')


if __name__ == '__main__':
  main()
